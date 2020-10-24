import time
from time import sleep
import threading
from selenium import webdriver
from threading import Thread
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as E
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.proxy import Proxy, ProxyType
from difflib import SequenceMatcher as sm
import tkinter as tk
from tkinter import Button
from openpyxl import load_workbook
import random
import string





class bot:
               
                

                def __init__(self, urlz, timez, browsername, proxy):

                                        
                                        self.time =timez
                                        self.url = urlz
                                        self.tabs = 0
                                        self.browsername = browsername
                                        self.PROXY = str(proxy)
                                        PROXY= self.PROXY
                                        print("proxy====="+str(self.PROXY))
                                        

                                        
                                        global urlfinal
                                        global timee
                                        timee=time
                                        urlfinal = urlz
                                        print(urlfinal)
                                        chrome_options= webdriver.ChromeOptions()




                                        

                                        prox = Proxy()
                                        prox.proxy_type = ProxyType.MANUAL
                                        prox.http_proxy = PROXY
                                        prox.ssl_proxy = PROXY
                                        prox.ftp_proxy = PROXY
                                        capabilities = webdriver.DesiredCapabilities.CHROME
                                        prox.add_to_capabilities(capabilities)

                                       # chrome_options = WebDriverWait.ChromeOptions()


                                        chrome_options.add_argument("--proxy-server='direct://'")
                                        chrome_options.add_argument("--proxy-bypass-list=*")
                                        #chrome_options.add_argument('--proxy-server=%s' % PROXY)
                                       # chrome_options.add_argument('--proxy-server=http://%s' % PROXY)

                                        chrome_options.add_experimental_option("excludeSwitches", ['enable-automation'])


                                        chrome_options.add_argument("--start-maximized")
                                        chrome_options.add_argument("ignore-certificate-errors")
                                        chrome_options.add_argument("--disable-notifications")
                                        chrome_options.add_argument("--no-sandbox")
                                        chrome_options.add_argument(" --disable-gpu")
                                        chrome_options.add_argument("--disable-dev-shm-usage")
                                        chrome_options.add_argument(f'user-agent={self.browsername}')
                                        self.driver= webdriver.Chrome(executable_path='chromedriver.exe', options=chrome_options )#,desired_capabilities=capabilities )

                                        

                def playlink_mobile(self):
                        try:
                                sleep(1)
                                playbutton= driver.find_element_by_class_name("ytp-cued-thumbnail-overlay").find_element_by_class_name("ytp-large-play-button")

                        except:
                                self.playlink()
                        action = ActionChains(driver)
                        action.move_to_element(playbutton)
                        action.click(on_element = playbutton)
                        action.perform()
                
                def playlink(self):
                        driver= self.driver
                        try:
                            setting = driver.find_element_by_class_name("ytp-right-controls").find_element_by_class_name("ytp-settings-button")
                        except:
                            sleep(1)
                            print("setting not found yetttt")
                            self.playlink()
                        
                        action = ActionChains(driver)
                        action.move_to_element(setting)
                        action.click(on_element = setting)
                        action.move_to_element(setting)
                        action.send_keys(Keys.DOWN)
                        action.move_to_element(setting)
                        action.send_keys(Keys.ENTER)
                        action.move_to_element(setting)
                        action.send_keys(Keys.UP)
                        action.move_to_element(setting)
                        action.send_keys(Keys.UP)
                        action.move_to_element(setting)
                        action.send_keys(Keys.UP)
                        action.move_to_element(setting)
                        action.send_keys(Keys.ENTER)
                        action.move_to_element(setting)
                        action.send_keys(Keys.DOWN)
                        action.move_to_element(setting)
                        action.send_keys(Keys.DOWN)
                        action.move_to_element(setting)
                        action.send_keys(Keys.ENTER)
                        action.move_to_element(setting)
                        action.send_keys(Keys.UP)
                        action.move_to_element(setting)
                        action.send_keys(Keys.ENTER)
                        action.move_to_element(setting)
                        try: 
                                action.perform()
                                action.click()
                        except:
                                sleep(5)
                                self.playlink()
                        play = driver.find_element_by_class_name("ytp-left-controls").find_element_by_class_name("ytp-play-button")   
                        aa = play.get_attribute("aria-label") 
                        if sm(None, a=aa,b="Play (k)").ratio() == 1:
                                    action1 = ActionChains(driver)
                                    action1.move_to_element(play)
                                    action1.click(on_element = play)
                                    action1.perform()

                        else:
                                        print("already playing")

                                                                       



                def search_for_things(self, url):
                    
                             driver=self.driver
                             def searchh():
                                    try:
                                        searchbox=driver.find_element_by_class_name("ytd-searchbox").find_element_by_id("search")
                                        
                                    except:
                                        sleep(1)
                                        searchh()
                             searchh()       

                             searchbox=driver.find_element_by_class_name("ytd-searchbox")
                            
                             action3 = ActionChains(driver)
                             action3.move_to_element(searchbox)
                             action3.click(on_element = searchbox)
                             action3.send_keys(url, Keys.ENTER)

                             action3.perform()
                        
                             sleep(1)
                             def fff():
                            
                                        try:
                                            
                                                videoclick = driver.find_elements_by_tag_name("ytd-video-renderer")
                                                print("50%")
                                                m= videoclick[0].find_element_by_id("video-title")
                                                print("75%")
                                                return m
                                            
                                                
                                
                                        except :
                                                print("didnt find")
                                                sleep(1)
                                                fff()
                                            
                                                   
                             videoclick = fff()
                             
                             action4 = ActionChains(driver)
                             action4.move_to_element(videoclick)
                             action4.click(on_element = videoclick)
                             action4.perform()
                             sleep(2)
                             self.playlink()
                               
                            

                        



                def stop_driver(self):
                                      driver= self.driver      
                                      for handle in range(self.tabs):
                                                print("time finished")
                                                
                                                driver.switch_to_window(driver.window_handles[0])
                                                driver.close()
                                                try:
                                                    driver.switch_to_window(driver.window_handles[-1])
                                                except:
                                                     driver.switch_to_window(driver.window_handles[0])
                                                     pass
                                                self.tabs=self.tabs-1
                                                print("tabs "+str(self.tabs))
                                                break
                                            
                                                
                                                                                                    
                       
                def start_chromedrivers(self):
        
                                timee = self.time       
                                self.driver.get(self.url)
                                self.playlink()
                                #t1 = threading.Timer(int(timee)*60, self.stop_driver)
                                #t1.start()
                                print(self.driver.execute_script("return document['$cdc_asdjflasutopfhvcZLmcfl_']"))

                              #  print(timee)
                               # print("tabs "+str(self.tabs))
                                
        
                                
                                
                                
                                                


              
                def new_tab(self, urlzz, time2):
                    print(urlzz)
                    url2=urlzz
                    print(url2)
                    driver = self.driver
                    driver.execute_script("window.open('');")
                    for handle in reversed(driver.window_handles):
                        driver.switch_to_window(handle)
                        driver.get(url2)
                        break
                    self.tabs= self.tabs+1
                    print(time2)
                    print("tabs "+str(self.tabs))
                    self.playlink()
                  #  t = threading.Timer(int(time2)*60, self.stop_driver)
                   # t.start()
        
                    
                
                    print("timer started")
                def url_entered():
                            window.withdraw()
                            lbl_temp = tk.Label(window, text=nmbrofbrowser.get())
                            lbl_temp.grid(row=1, column=2, sticky="nsew")
                            print(nmbrofbrowser.get())
                            j = int(nmbrofbrowser.get())
                            j=j-1
                            print(url.get())
                            url2.grid(row=5, column=2, sticky="nsew")
                            urls = url.get()
                            url.destroy()
                            labelnothn=tk.Label(window, text= "enter another url ").grid(row=4, column=2, sticky="nsew")

                            start_chromedrivers(j, urls)
                            url_button2=  tk.Button(window, text="enter another url", command=new_tab).grid(row=6, column=2, sticky="nsew")
                 
                            window.deiconify()
                                      
                        
                            
               


if __name__ == '__main__':
                
        
    window = tk.Tk()
    window.rowconfigure(8, minsize=50, weight=1)
    window.columnconfigure(3, minsize=50, weight=5)



    nmbrofbrowser = tk.Entry( width=10)
    nmbrofbrowser.grid(row=1, column=2, sticky="nsew")
    url = tk.Entry( window, width=20, )
    url.grid(row=5, column=2, sticky="nsew")
    url2 = tk.Entry( window, width=20, )
    labelnothn2=tk.Label(window, text= "        ").grid(row=7, column=2, sticky="nsew")
    time=tk.Entry(window, width=10 )
    time.grid(row=7, column=2, sticky="nsew")



    obj=[]
    
                    


    def newlinkbutton():
                window.withdraw()
                url3=url2.get()
                timee=int(time.get())
                for browser in obj:
                        
                        t1= threading.Thread(target=browser.new_tab, args=(url3,timee))
                        t1.start()

                    
                window.deiconify()


   



    def button_press():

            workbook = load_workbook(filename="user_agents.xlsx")
            workbook.sheetnames
            ['Sheet 1']
            sheet = workbook.active
            sheet
            useragentsssss =  []
            PROXIES =  []
            bad_chars= ["(","'",",",")"]
    
            for value in sheet.iter_rows(min_row=2,max_row=100,
                                         max_col=1, min_col =1,
                                          values_only=True):
                            useragentsssss.append(value)

            for valu in sheet.iter_rows(min_row=2,max_row=100,
                                         max_col=2, min_col =2,
                                          values_only=True):
                            try:
                                valu = ''.join((filter(lambda i: i not in bad_chars, valu)))
                            except TypeError:
                                break
                                


                            
                            PROXIES.append(valu)
                            
                            print(valu)
                            
            print("lwnth poxy="+str(len(PROXIES)))








        
            window.withdraw()
            lbl_temp = tk.Label(window, text=nmbrofbrowser.get())
            lbl_temp.grid(row=1, column=2, sticky="nsew")
            url2.grid(row=5, column=2, sticky="nsew")
            labelnothn=tk.Label(window, text= "enter another url ").grid(row=4, column=2, sticky="nsew")
            url_button2=  tk.Button(window, text="enter another url", command=newlinkbutton).grid(row=6, column=2, sticky="nsew")
                     
                                   
            i=0
            url1=str(url.get())
            print(url1)
            url.destroy()

            time1 = int(time.get())
            print(time1)
            j=int(nmbrofbrowser.get())
            while i < j:
                     obj.append(i)
                     obj[i] = bot(url1, time1, useragentsssss[i], PROXIES[i] )
                     i = i+1
            processsess=[]
            print("size="+str(len(obj)))
            for things in obj:
              t=  threading.Thread(target=things.start_chromedrivers)
              t.start()

                    
            window.deiconify()
           
                    
                    
                    

    labelnothn=tk.Label(window, text= "number of browser")
    labelnothn.grid(row=0, column=2, sticky="nsew")


    labelnothn3=tk.Label(window, text= "        ").grid(row=3, column=2, sticky="nsew")


    labelnothn=tk.Label(window, text= "enter url ").grid(row=4, column=2, sticky="nsew")

    btn_increase = tk.Label(window, text="__________________________________________")
    btn_increase.grid(row=2, column=2, sticky="nsew")


    url_button=  tk.Button(window, text="enter url", command= button_press).grid(row=6, column=2, sticky="nsew")


    window.mainloop()






                    

